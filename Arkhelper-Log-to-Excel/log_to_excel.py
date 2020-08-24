from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

def start():
    log_file = 'ArknightsAutoHelper.log'
    encoding = 'utf-8'
    mode = 'r'

    read_log(log_file, mode, encoding)

def read_log(log_file, mode, encoding):
    with open(log_file, mode=mode, encoding=encoding) as f:
        data = f.readlines()
        log_to_excel(data)

def log_to_excel(data):

    wb = Workbook()
    ws = wb.active
    ws.title = '战斗日志'
    
    ws['A1'] = '时间'
    ws['B1'] = '位置'
    ws['C1'] = '状态'
    ws['D1'] = '详情'
    
    for line in data:
        row = line.split(' ! ')     # 以 '!' 分隔各类信息
        if len(row) == 4:            # 排除异常，只保留正常记录数据
            ws.append(row)

    beautify_excel(ws, data)
    
    wb.save('ArknightsAutoHelper.xlsx')
    print('保存成功！')

def beautify_excel(ws, data):

    # 首行加粗、居中
    for cell in ws[1]:                  
        cell.font = Font(bold=True, size=20)
        cell.alignment = CustomStyle.center_align

    # 【时间】列居中、调整宽度
    for cell in ws['A']:
        cell.alignment = CustomStyle.center_align
    ws.column_dimensions['A'].width = 25

    # 【状态】列
    for cell in ws['C']:
        cell.alignment = CustomStyle.center_align
        state = cell.value  # 读取状态
        if (state == "INFO"):
            cell.font = CustomStyle.white_bold_font
            cell.fill = PatternFill(fill_type='solid', fgColor='00B0F0')    # 蓝色背景

    stone_cnt = 0   # 计算源石消耗数量
    level_up = 0    # 升级次数
    total_prds = dict()  # 计算日志中所有的产物获得数
    cur_prds = dict()   # 本轮产出
    chapters = dict()   # 关卡计算

    for tup in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        cell = tup[0]
        text = cell.value
        if text.startswith('掉'):       # 【掉落识别结果：...】，表示【本次】刷图的掉落
            cell.font = CustomStyle.white_bold_font
            cell.fill = PatternFill(fill_type='solid', fgColor='008E40')    # 深绿色背景

        elif text.startswith('目'):     # 【目前已获得：...】，表示【本轮】累积获得的产出
            # 样式设置
            cell.font = CustomStyle.white_bold_font
            cell.fill = PatternFill(fill_type='solid', fgColor='DA9694')    # 粉色背景
            # 产出分析
            prds_list = text[6:].split(', ')
            for prd in prds_list:
                p = prd.split('x')
                prd_name = p[0]
                prd_num = int(p[-1])
                cur_prds[prd_name] = prd_num    

        elif text.startswith('碎'):     # 【碎石回复理智】，表示已消耗源石
            cell.alignment = CustomStyle.center_align
            cell.fill = PatternFill(fill_type='solid', fgColor='FFC000')    # 橙黄色背景
            cell.font = CustomStyle.bold_font
            stone_cnt += 1

        elif text.startswith('简'):     # 【简略模块None结束】，刷图结束标志
            cell.alignment = CustomStyle.center_align
            cell.fill = PatternFill(fill_type='solid', fgColor='000000')    # 黑色背景
            cell.font = CustomStyle.white_bold_font
            update_prd(total_prds, cur_prds)    # 一轮结束，更新总产出

        elif text.startswith('关'):     # 【关闭升级提示】，表示在刷图过程中升级了
            level_up += 1

        elif text.startswith('当前画面关卡'):   #【当前画面关卡】，计算关卡通过次数
            cpt = text[7:-1]    # 提取关卡名称，去除 \n
            if cpt not in chapters:
                chapters[cpt] = 1
            else:
                chapters[cpt] += 1
                        
        cell.alignment = Alignment(wrapText=True)

    ws.column_dimensions['D'].width = 50        # 【详细】列宽调整

    data_sum(ws, data, stone_cnt, total_prds, level_up, chapters) # 总结

def data_sum(ws, data, stone_cnt, total_prds, level_up, chapters):
    idx = 2
    merge_coordinates = set()
    
    start_time, end_time = data[0].split(' ! ')[0], data[-1].split(' ! ')[0]
    start_time_text_cell, end_time_text_cell = ws[f'F{idx}'], ws[f'F{idx+1}']
    start_time_cell, end_time_cell = ws[f'G{idx}'], ws[f'G{idx+1}']
    start_time_text_cell.value, end_time_text_cell.value = '日志记录开始时间：', '日志记录结束时间：'
    start_time_cell.value, end_time_cell.value = start_time, end_time
    start_time_cell.font = end_time_cell.font = Font(bold=True, color='FF0000', size=22)
    idx += 2
    
    stone_text_cell = ws[f'F{idx}']
    stone_cnt_cell = ws[f'G{idx}']
    stone_text_cell.value = '使用至纯源石数量：'
    stone_text_cell.fill = PatternFill(fill_type='solid', fgColor='EAD492')
    stone_text_cell.font = Font(bold=True, size=22)
    stone_cnt_cell.value = stone_cnt
    stone_cnt_cell.font = Font(bold=True, color='EAD492', size=22)
    stone_cnt_cell.alignment = CustomStyle.center_align
    idx += 1
    
    level_text_cell, level_cell = ws[f'F{idx}'], ws[f'G{idx}']
    level_text_cell.value = '刷图期间提升等级：'
    level_cell.value = level_up
    level_cell.alignment = CustomStyle.center_align
    level_cell.font = Font(bold=True, size=22, color='D07EF1')
    idx += 2

    ws.merge_cells(f'F{idx}:G{idx}')
    merge_cell = ws[f'F{idx}']
    merge_coordinates.add(merge_cell)
    merge_cell.value = '掉落统计'
    merge_cell.font = Font(bold=True, size=22)
    idx += 1
    
    for name, num in total_prds.items():
        name_cell, num_cell = ws['F'+str(idx)], ws['G'+str(idx)]
        name_cell.value, num_cell.value = name, num
        num_cell.alignment = CustomStyle.center_align
        if(name == '龙门币'):
            num_cell.font = Font(bold=True, size=22, color='22BBFF')
        elif(name == '声望'):
            num_cell.font = Font(bold=True, size=22, color='9BD145')
        elif(name == '合成玉'):
            num_cell.font = Font(bold=True, size=22, color='B31713')
        else:
            num_cell.font = Font(bold=True, size=22)
        name_cell.font = Font(size=20)
        idx += 1

    idx += 1
    ws.merge_cells(f'F{idx}:G{idx}')
    merge_cell = ws[f'F{idx}']
    merge_coordinates.add(merge_cell)
    merge_cell.value = '完成关卡统计'
    merge_cell.font = Font(bold=True, size=22)
    idx += 1

    for cpt, num in chapters.items():
        name_cell, num_cell = ws[f'F{idx}'], ws[f'G{idx}']
        name_cell.value, num_cell.value = cpt, num
        name_cell.font = num_cell.font = Font(bold=True, size=20)
        num_cell.alignment = CustomStyle.center_align
        idx += 1
        
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['F'].width = 40
    
    # 【文本】列水平靠右，垂直居中
    for cell in ws['F']:
        if cell in merge_coordinates:
            cell.alignment = CustomStyle.center_align
        else:
            cell.alignment = CustomStyle.right_center_align
        
def update_prd(total, cur):
    for k, v in cur.items():
        if k not in total:
            total[k] = v
        else:
            total[k] += v
    cur.clear()

class CustomStyle():
    center_align = Alignment(horizontal='center', vertical='center')
    right_center_align = Alignment(horizontal='right', vertical='center')
    bold_font = Font(bold=True)
    white_bold_font = Font(bold=True, color='FFFFFF')

if __name__ == "__main__":
    start()
